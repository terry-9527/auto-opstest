import configparser
import yaml
import os
from openpyxl import load_workbook
from utils.handle_path import ini_dir,testdata_dir,testcase_dir


ROOT_PATH = str(os.path.abspath(os.getcwd()).split('yunweixitong')[0]) + "yunweixitong"



class readData():
    """
    封账读取数据的方法
    """

    def read_config(self, section, option, filename=None):
        """
        读取配置文件，返回对应的配置值
        :param section:
        :param option:
        :param filename:
        :return:
        """
        if not filename:
            config_path = os.path.join(ini_dir, 'config.ini')
        else:
            config_path = os.path.join(ini_dir, filename)
        config = configparser.RawConfigParser()
        config.read(config_path, encoding='utf-8')
        result = config.get(section, option)
        return result

    def write_config(self, section, option, value=None, filename=None):
        if not filename:
            config_path = os.path.join(ini_dir, 'config.ini')
        else:
            config_path = os.path.join(ini_dir, filename)
        config = configparser.RawConfigParser()
        config.read(config_path, encoding="utf-8")
        config.set(section, option, value)
        with open(config_path,'w',encoding="utf-8") as file:
            config.write(file)

    def read_yaml(self, filename):
        """
        读取yaml用例文件
        :param filename:
        :return:
        """
        file_path = os.path.join(testcase_dir, filename)
        with open(file_path, 'r', encoding='utf-8') as f:
            yaml_data = yaml.load(f, Loader=yaml.FullLoader)  # 读取yaml文件内容，返回dict数据
            case_data = []
            for case in yaml_data.values():
                case_data.append(case)
            f.close()
        return case_data

    def read_excel(self, sheetname, filename):
        """
        读取Excel用例文件
        :param sheetname: 表名
        :param file: 文件名
        :return: 返回一个列表
        """
        file_path = os.path.join(testdata_dir, filename)
        workbook = load_workbook(file_path)
        sheet = workbook[sheetname]  # 执行使用哪个工作表,根据传进来的表名称决定读取对应的数据
        rows = sheet.rows  # 取出所有行的数据
        cases = []
        num = 1

        # 遍历每一行的数据
        for row in rows:
            list1 = []
            cycle_number = 1
            # 取出第一行执行用例execute字段
            if num == 1:
                for i in range(len(row)):
                    if row[i].value == "execute":
                        self.index = i
                        num += 1
            # 读取需要执行的用例，execute列值为True的代表需要执行
            if row[self.index].value == "True":  # 判断该条用例是否需要执行,execute列的值为True则执行，进行数据读取
                for col in row:
                    if not col.value:  # 处理空单元格，直接添加None
                        list1.append(col.value)
                    elif col.value.startswith('{') and col.value.endswith('}'):
                        # if cycle_number == 3:
                        #     list1.append(tuple(eval(col.value).items()))
                        # elif cycle_number == 2:
                        #     list1.append(tuple(eval(col.value).values()))
                        # else:
                        #     cycle_number += 1
                        list1.append(eval(col.value))  # 获取当前行每一个单元格，单元格的内容需要用eval，把str类型转化为dict类型
                    else:
                        list1.append(col.value)
                cases.append(list1)
        return cases

    # Excel写入数据 写入测试结果、写入失败的原因
    def write_excel(self, filename, sheetname, case_id=None, testresult=None, reason=None):
        file_path = os.path.join(testcase_dir, filename)
        workbook = load_workbook(file_path)
        # sheets = workbook.sheetnames  # 获取所有工作表名称
        sheet = workbook[sheetname]  # 执行使用哪个工作表
        cols = sheet.columns
        col = [col for col in cols]
        list1 = []
        for row in col[0]:
            list1.append(row.value)
        i = 0
        for id in list1:
            if case_id == id:
                i += 1
                break
            i += 1
        sheet.cell(i, sheet.max_column - 1).value = testresult  # 写入测试的结果
        sheet.cell(i, sheet.max_column).value = reason  # 写入失败的原因
        workbook.save(file_path)

    def read_sqls(self, filename):
        file_path = os.path.join(testdata_dir, filename)
        sqls = []
        with open(file_path, 'r', encoding='utf-8') as stream:
            for line in stream.readlines():
                if not len(line.strip()) or line.startswith('--'):
                    continue
                sqls.append(line.strip())
        return sqls


if __name__ == '__main__':
    read = readData()
    # sqls = read.read_sqls('systemsetting.txt')
    # print(sqls)
    datas = read.read_excel("添加模板类型", "commandinfo.xlsx")
    print(datas[0][0])

    # if datas[1][0] in ["edit-clusterinfo-002", "edit-clusterinfo-003"]:
    #     print("1111")
    # # data = read.read_config("publicCloud", "ssh_port", "config.ini")
    # # print(data)
    # data = read.read_config("publicCloud", "host", filename="config12.ini")
    # print(data)
